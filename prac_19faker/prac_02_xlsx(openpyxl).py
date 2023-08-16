from faker import Faker
from openpyxl import load_workbook
"""
        将faker数据写入excel文件(使用openpyxl库)
"""
fake = Faker(locale='zh-CN')  # 创建一个使用中文数据的Faker对象

# 打开指定的Excel文件
wb = load_workbook('inner_employee.xlsx')

# 获取要操作的工作表
ws = wb['Sheet1']  # 假设要操作的工作表名为"Sheet1"

# 指定要写入数据的行和列
row_index = 4  # 假设要从第4行开始写入数据
col_index = 1  # 假设要从第1列开始写入数据

# 生成并写入随机数据
for _ in range(10):  # 假设需要生成10行数据
    name = fake.unique.name()                            # 中文成员姓名
    english_name = Faker(locale="en_US").unique.name()   # 英文签名
    number = fake.unique.pyint()                         # 编号 这里用随机Int数字
    phone_number = fake.unique.phone_number()            # 手机号
    email = fake.unique.email()                          # 邮箱

    print(f"{name}  {english_name}    {number}  {phone_number}  {email}")

    # 将数据写入指定的单元格
    ws.cell(row=row_index, column=col_index, value=name)
    ws.cell(row=row_index, column=col_index + 1, value=english_name)
    ws.cell(row=row_index, column=col_index + 2, value=number)
    ws.cell(row=row_index, column=col_index + 4, value=phone_number)
    ws.cell(row=row_index, column=col_index + 5, value=email)
    ws.cell(row=row_index, column=col_index + 6, value="众畅科技")
    ws.cell(row=row_index, column=col_index + 7, value="系统管理员；印章管理员；流程管理员；模板管理员")
    ws.cell(row=row_index, column=col_index + 9, value="qiyuesuo#2020")
    row_index += 1

# 保存修改后的Excel文件
wb.save('inner_employee.xlsx')
